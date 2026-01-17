"""Сервис генерации отчетов (PDF, Excel с форматированием)"""
from typing import List, Dict, Any, Optional
from pathlib import Path
from app.utils.logger import logger

# PDF генерация
PDF_AVAILABLE = False
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    PDF_AVAILABLE = True
except ImportError:
    pass


class ReportGenerator:
    """Сервис для генерации отчетов"""
    
    def __init__(self):
        self.pdf_available = PDF_AVAILABLE
    
    def generate_pdf_report(
        self,
        data: List[Dict[str, Any]],
        output_path: str,
        title: str = "Providers Report",
        include_statistics: bool = True
    ) -> Optional[str]:
        """
        Сгенерировать PDF отчет
        
        Args:
            data: Список данных провайдеров
            output_path: Путь для сохранения PDF
            title: Заголовок отчета
            include_statistics: Включить статистику
        
        Returns:
            Путь к созданному файлу или None при ошибке
        """
        if not self.pdf_available:
            logger.warning("PDF generation not available (reportlab not installed)")
            return None
        
        try:
            doc = SimpleDocTemplate(output_path, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()
            
            # Заголовок
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1a1a1a'),
                spaceAfter=30,
                alignment=1  # Center
            )
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 0.2*inch))
            
            # Статистика (если нужно)
            if include_statistics and data:
                stats_text = f"<b>Всего записей:</b> {len(data)}"
                story.append(Paragraph(stats_text, styles['Normal']))
                story.append(Spacer(1, 0.2*inch))
            
            # Таблица данных
            if data:
                # Заголовки
                headers = ['ID', 'Merchant', 'Provider Domain', 'Account Type', 'Payment Method']
                table_data = [headers]
                
                # Данные (первые 100 для читаемости)
                for item in data[:100]:
                    row = [
                        str(item.get('id', '')),
                        item.get('merchant', ''),
                        item.get('provider_domain', ''),
                        item.get('account_type', ''),
                        item.get('payment_method', '')
                    ]
                    table_data.append(row)
                
                # Создание таблицы
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a90e2')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                ]))
                
                story.append(table)
                
                if len(data) > 100:
                    story.append(Spacer(1, 0.2*inch))
                    story.append(Paragraph(
                        f"<i>Показано первых 100 из {len(data)} записей</i>",
                        styles['Normal']
                    ))
            
            # Генерация PDF
            doc.build(story)
            
            logger.info(f"PDF report generated: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"Error generating PDF report: {e}", exc_info=True)
            return None
    
    def generate_excel_report_formatted(
        self,
        data: List[Dict[str, Any]],
        output_path: str,
        title: str = "Providers Report"
    ) -> Optional[str]:
        """
        Сгенерировать Excel отчет с форматированием
        
        Args:
            data: Список данных провайдеров
            output_path: Путь для сохранения Excel
            title: Заголовок отчета
        
        Returns:
            Путь к созданному файлу или None при ошибке
        """
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # Создаем DataFrame
            df = pd.DataFrame(data)
            
            if df.empty:
                logger.warning("No data to export")
                return None
            
            # Сохраняем в Excel
            df.to_excel(output_path, index=False, engine='openpyxl')
            
            # Применяем форматирование
            wb = load_workbook(output_path)
            ws = wb.active
            
            # Заголовки - жирный шрифт, цветной фон
            header_fill = PatternFill(start_color="4a90e2", end_color="4a90e2", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Автоматическая ширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Заморозить первую строку
            ws.freeze_panes = "A2"
            
            wb.save(output_path)
            
            logger.info(f"Excel report generated: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}", exc_info=True)
            return None


# Глобальный экземпляр
_report_generator: Optional[ReportGenerator] = None


def get_report_generator() -> ReportGenerator:
    """Получить глобальный report generator"""
    global _report_generator
    if _report_generator is None:
        _report_generator = ReportGenerator()
    return _report_generator
